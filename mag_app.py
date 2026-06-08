import streamlit as st
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import io
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

# ══════════════════════════════════════════════════════════
#  НАЛАШТУВАННЯ СТОРІНКИ
# ══════════════════════════════════════════════════════════
st.set_page_config(
    page_title="MagnetModel Pro",
    page_icon="🌿",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ══════════════════════════════════════════════════════════
#  МАТЕМАТИЧНА МОДЕЛЬ
# ══════════════════════════════════════════════════════════

def _design_matrix(x1: np.ndarray, x2: np.ndarray) -> np.ndarray:
    """Матриця ознак: 6 базисних функцій."""
    return np.column_stack([
        x1 * (x2**2),    # F1
        x1 * x2,         # F2
        x2**2,           # F3
        x1,              # F4
        x2,              # F5
        np.ones_like(x1) # F6
    ])

def ridge_regression(A: np.ndarray, y: np.ndarray, alpha: float = 0.01) -> np.ndarray:
    """Стабільний Ridge МНК: θ = (AᵀA + αI)⁻¹ Aᵀy."""
    I = np.eye(A.shape[1])
    return np.linalg.inv(A.T @ A + alpha * I) @ A.T @ y

def evaluate_model(x1, x2, y_true, coeffs):
    """Оцінка моделі — R² та RMSE."""
    A        = _design_matrix(x1, x2)
    y_pred   = A @ coeffs
    residuals = y_true - y_pred
    ss_res   = np.sum(residuals**2)
    ss_tot   = np.sum((y_true - np.mean(y_true))**2)
    r2       = 1 - ss_res / ss_tot if ss_tot > 0 else 0.0
    rmse     = np.sqrt(np.mean(residuals**2))
    return y_pred, residuals, r2, rmse

def calculate_Y_physical(x1: float, x2: float, coeffs: np.ndarray) -> float:
    """Прогноз одного значення."""
    F1, F2, F3, F4, F5, F6 = coeffs
    return (
        F1 * x1 * (x2**2) +
        F2 * x1 * x2 +
        F3 * (x2**2) +
        F4 * x1 +
        F5 * x2 +
        F6
    )

def to_float_array(series: pd.Series) -> np.ndarray:
    """Безпечна конвертація серії в float64."""
    return pd.to_numeric(series, errors='raise').to_numpy(dtype=np.float64)

# ══════════════════════════════════════════════════════════
#  ЗБЕРЕЖЕННЯ СТАНУ ЗАСТОСУНКУ
# ══════════════════════════════════════════════════════════
if "train_df" not in st.session_state:
    st.session_state.train_df = pd.DataFrame(columns=["X1", "X2", "Y"])
if "pred_df" not in st.session_state:
    st.session_state.pred_df = pd.DataFrame(columns=["X1", "X2", "Y_predicted"])
if "coeffs" not in st.session_state:
    st.session_state.coeffs = None
if "metrics" not in st.session_state:
    st.session_state.metrics = {"R2": "—", "RMSE": "—"}

# ══════════════════════════════════════════════════════════
#  САЙДБАР
# ══════════════════════════════════════════════════════════
with st.sidebar:
    st.title("🌿 MagneticIE Pro")
    st.caption("Ridge Regression Engine · Магнітна активація")
    st.markdown("---")

    # Секція 1: Введення даних вручну
    st.subheader("⬡ Введення даних")
    inp_x1 = st.number_input("X₁ — Концентрація", value=0.0, step=0.1, format="%.4f")
    inp_x2 = st.number_input("X₂ — Напруженість", value=0.0, step=1.0, format="%.4f")
    inp_y  = st.number_input("Y — Вихідний параметр (для навч.)", value=0.0, step=0.1, format="%.4f")

    if st.button("➕ Додати рядок до навчання", use_container_width=True):
        new_row = pd.DataFrame([{"X1": float(inp_x1), "X2": float(inp_x2), "Y": float(inp_y)}])
        st.session_state.train_df = pd.concat(
            [st.session_state.train_df, new_row], ignore_index=True
        )
        st.success("Рядок успішно додано!")
        st.rerun()

    # Секція 2: Завантаження файлу
    st.subheader("📂 Імпорт файлу")
    uploaded_file = st.file_uploader("Оберіть файл Excel або CSV", type=["xlsx", "xls", "csv"])
    if uploaded_file is not None:
        try:
            if uploaded_file.name.endswith('.csv'):
                df_file = pd.read_csv(uploaded_file)
            else:
                df_file = pd.read_excel(uploaded_file)

            if len(df_file.columns) >= 3:
                df_file = df_file.iloc[:, :3].copy()
                df_file.columns = ["X1", "X2", "Y"]
                # Конвертуємо одразу при завантаженні
                for col in ["X1", "X2", "Y"]:
                    df_file[col] = pd.to_numeric(df_file[col], errors='coerce')
                df_file.dropna(inplace=True)
                st.session_state.train_df = pd.concat(
                    [st.session_state.train_df, df_file], ignore_index=True
                )
                st.success(f"Завантажено {len(df_file)} рядків!")
                st.rerun()
            else:
                st.error("Файл повинен містити мінімум 3 колонки (X1, X2, Y)")
        except Exception as e:
            st.error(f"Помилка читання файлу: {e}")

    # Секція 3: Керування моделлю
    st.subheader("⚙️ Модель")
    if st.button("⚡ Навчити модель", type="primary", use_container_width=True):
        if len(st.session_state.train_df) < 6:
            st.warning(f"Потрібно ≥6 тренувальних точок (зараз: {len(st.session_state.train_df)})")
        else:
            try:
                x1 = to_float_array(st.session_state.train_df["X1"])
                x2 = to_float_array(st.session_state.train_df["X2"])
                y  = to_float_array(st.session_state.train_df["Y"])

                A = _design_matrix(x1, x2)
                c = ridge_regression(A, y, alpha=0.01)
                _, _, r2, rmse = evaluate_model(x1, x2, y, c)

                st.session_state.coeffs = c
                st.session_state.metrics["R2"]   = f"{r2:.4f}"
                st.session_state.metrics["RMSE"] = f"{rmse:.4f}"
                st.success("Модель успішно навчено!")
                st.rerun()
            except ValueError:
                st.error("Помилка даних: Перевірте, чи таблиця не містить тексту або порожніх клітинок!")
            except np.linalg.LinAlgError as e:
                st.error(f"Помилка лінійної алгебри: {e}")
            except Exception as e:
                st.error(f"Непередбачена помилка: {e}")

    # Секція 4: Прогноз
    st.subheader("◎ Прогноз")
    if st.button("🔮 Зробити прогноз", use_container_width=True):
        if st.session_state.coeffs is None:
            st.error("Спочатку навчіть модель!")
        else:
            y_pred_val = calculate_Y_physical(inp_x1, inp_x2, st.session_state.coeffs)
            new_pred = pd.DataFrame([{
                "X1": float(inp_x1),
                "X2": float(inp_x2),
                "Y_predicted": round(y_pred_val, 4)
            }])
            st.session_state.pred_df = pd.concat(
                [st.session_state.pred_df, new_pred], ignore_index=True
            )
            st.toast(f"Прогноз готовий: Y = {y_pred_val:.4f}", icon="🔮")
            st.rerun()

    st.markdown("---")
    if st.button("🗑 Очистити всі дані", type="secondary", use_container_width=True):
        st.session_state.train_df = pd.DataFrame(columns=["X1", "X2", "Y"])
        st.session_state.pred_df  = pd.DataFrame(columns=["X1", "X2", "Y_predicted"])
        st.session_state.coeffs   = None
        st.session_state.metrics  = {"R2": "—", "RMSE": "—"}
        st.toast("Всі дані скинуто", icon="🗑")
        st.rerun()

# ══════════════════════════════════════════════════════════
#  ГОЛОВНА ПАНЕЛЬ
# ══════════════════════════════════════════════════════════

col_m1, col_m2, col_m3, col_m4 = st.columns(4)
with col_m1:
    st.metric("R²", st.session_state.metrics["R2"])
with col_m2:
    st.metric("RMSE", st.session_state.metrics["RMSE"])
with col_m3:
    st.metric("Навч. точок", len(st.session_state.train_df))
with col_m4:
    st.metric("Прогнозів", len(st.session_state.pred_df))

st.markdown("### 📋 Коефіцієнти моделі")
if st.session_state.coeffs is not None:
    c = st.session_state.coeffs
    st.code(
        f"F₁={c[0]:.6e}   F₂={c[1]:.6e}   F₃={c[2]:.6e}\n"
        f"F₄={c[3]:.6e}   F₅={c[4]:.6e}   F₆={c[5]:.6e}",
        language="text"
    )
else:
    st.info("Модель не навчена")

tab_data, tab_plots = st.tabs(["📋 Дані застосунку", "📈 Графічний аналіз"])

# ── ВКЛАДКА 1: ДАНІ ──────────────────────────────────────
with tab_data:
    col_t1, col_t2 = st.columns(2)

    with col_t1:
        st.subheader("Дані для навчання моделі")
        st.dataframe(st.session_state.train_df, use_container_width=True, height=300)

    with col_t2:
        st.subheader("Результати прогнозів")
        st.dataframe(st.session_state.pred_df, use_container_width=True, height=300)

        if not st.session_state.pred_df.empty:
            buffer = io.BytesIO()
            wb = Workbook()
            ws = wb.active
            ws.title = "Прогнози"

            headers  = ["X1", "X2", "Y_predicted"]
            hdr_fill = PatternFill("solid", fgColor="2EE59D")
            hdr_font = Font(bold=True, color="0A1A12")
            thin     = Side(style="thin", color="1E4A30")
            brd      = Border(left=thin, right=thin, top=thin, bottom=thin)

            for ci, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=ci, value=h)
                cell.fill = hdr_fill
                cell.font = hdr_font
                cell.alignment = Alignment(horizontal="center")
                cell.border = brd
                ws.column_dimensions[cell.column_letter].width = 18

            for ri, row in enumerate(st.session_state.pred_df.itertuples(), 2):
                for ci, v in enumerate([row.X1, row.X2, row.Y_predicted], 1):
                    cell = ws.cell(row=ri, column=ci, value=v)
                    cell.alignment = Alignment(horizontal="center")
                    cell.border = brd

            wb.save(buffer)
            buffer.seek(0)

            st.download_button(
                label="💾 Скачати прогнози в Excel",
                data=buffer,
                file_name="mag_predictions.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )

# ── ВКЛАДКА 2: ГРАФІКИ ───────────────────────────────────
with tab_plots:
    if st.session_state.coeffs is None:
        st.info("Будь ласка, завантажте дані та навчіть модель для побудови графіків.")
    else:
        # ВИПРАВЛЕННЯ: безпечна конвертація даних у float64
        try:
            x1 = to_float_array(st.session_state.train_df["X1"])
            x2 = to_float_array(st.session_state.train_df["X2"])
            y  = to_float_array(st.session_state.train_df["Y"])
        except Exception as e:
            st.error(f"Помилка читання даних для графіків: {e}")
            st.stop()

        y_pred, residuals, r2, rmse = evaluate_model(x1, x2, y, st.session_state.coeffs)

        plot_type = st.radio(
            "Оберіть графік аналізу:",
            ["Exp vs Predicted (Адекватність)", "Аналіз залишків", "Прогнозні значення"],
            horizontal=True
        )

        fig, ax = plt.subplots(figsize=(10, 5))

        if plot_type == "Exp vs Predicted (Адекватність)":
            ax.scatter(y, y_pred, color="#2EE59D", s=80, zorder=5, alpha=0.9,
                       edgecolors="#5FFFC0", label="Точки")
            mn = min(y.min(), y_pred.min())
            mx = max(y.max(), y_pred.max())
            ax.plot([mn, mx], [mn, mx], color="#A3FF6F", lw=2, ls="--", label="Ідеал y=x")
            ax.set_xlabel("Експериментальні Y")
            ax.set_ylabel("Модельні Y")
            ax.set_title(f"Адекватність моделі — R²={r2:.4f}  RMSE={rmse:.4f}")
            ax.legend()

        elif plot_type == "Аналіз залишків":
            ax.scatter(y_pred, residuals, color="#2EE59D", s=80, alpha=0.9,
                       edgecolors="#5FFFC0")
            ax.axhline(0, color="#A3FF6F", lw=2, ls="--")
            ax.set_xlabel("Прогнозовані Y")
            ax.set_ylabel("Залишки (Y_exp − Y_pred)")
            ax.set_title("Аналіз залишків моделі")

        elif plot_type == "Прогнозні значення":
            if st.session_state.pred_df.empty:
                st.warning(
                    "Немає точок прогнозів для відображення кривих. "
                    "Будь ласка, зробіть кілька прогнозів у сайдбарі."
                )
            else:
                # ВИПРАВЛЕННЯ: конвертуємо pred_df колонки у float перед сортуванням
                df_p = st.session_state.pred_df.copy()
                df_p["X1"] = pd.to_numeric(df_p["X1"], errors='coerce')
                df_p["X2"] = pd.to_numeric(df_p["X2"], errors='coerce')
                df_p["Y_predicted"] = pd.to_numeric(df_p["Y_predicted"], errors='coerce')
                df_p.dropna(inplace=True)
                df_p = df_p.sort_values("X2")

                palette = ["#2EE59D", "#A3FF6F", "#5BC8FF", "#FFD166", "#FF5C7A"]

                # ВИПРАВЛЕННЯ: округлюємо X1 для коректного групування (уникаємо float-дублів)
                df_p["X1_rounded"] = df_p["X1"].round(6)
                unique_x1 = sorted(df_p["X1_rounded"].unique())

                for ci, x1_val in enumerate(unique_x1):
                    sub = df_p[df_p["X1_rounded"] == x1_val]
                    color = palette[ci % len(palette)]
                    ax.plot(sub["X2"], sub["Y_predicted"],
                            marker="o", color=color, linewidth=2.2,
                            label=f"X₁={x1_val:.4g}")
                ax.set_xlabel("X₂")
                ax.set_ylabel("Y (прогноз)")
                ax.set_title("Прогнозні значення моделі")
                ax.legend()

        fig.tight_layout()
        st.pyplot(fig)
